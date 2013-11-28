# -*- coding: gbk -*-

import urllib2
from openpyxl import load_workbook

filename = raw_input("������Ҫ������ļ�����ֱ�Ӱ��»س�ΪĬ���ļ���(IP�ع�.xlsx):")
open("result.txt", "w").write("")

if not filename:
    filename = 'IP�ع�.xlsx'

wb = load_workbook(filename = 'IP�ع�.xlsx')
sheet_name = wb.get_sheet_names()[0]
sheet = wb.get_sheet_by_name(sheet_name)

n = 1
while True:
    n += 1
    if not sheet.cell('A' + str(n)).value:
        break
    ip = str(sheet.cell('A' + str(n)).value)
    print ip,
    url = "http://int.dpool.sina.com.cn/iplookup/iplookup.php?ip=%s" % ip
    r = urllib2.urlopen(url).read()
    rs = r.split()[3:]
    address = " ".join(rs)
    print address
    sheet.cell('B' + str(n)).value = address.decode("gbk")
    
    open("result.txt", "a").write(ip + "," + address + "\n")

wb.save('result.xlsx')
raw_input("���! ��������� result.xlsx �ļ�")
